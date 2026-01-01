"""
Smart Classifier v3 - Multi-Model Classification

Uses different models based on file type:
- Excel/CSV: Qwen (text-based)
- PDF: Llama 3.2 Vision (image-based)

Features:
- Quick-classify mode (type only, fast skip decision)
- Two-phase full classification
- Retry logic for JSON errors
- Column validation
"""

import sys
import json
import logging
import base64
import re
from pathlib import Path
from io import StringIO, BytesIO
import time
from pypdf import PdfReader

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

from scripts.ingestion_lib import call_ollama, parse_json_garbage

# CONFIG
QWEN_MODEL = "qwen2.5:14b"
VISION_MODEL = "llama3.2-vision:11b"
QUICK_SAMPLE_ROWS = 5  # Minimal sample for quick-classify (avoids timeout)
FULL_SAMPLE_ROWS = 10  # Full sample for detailed mapping
MAX_RETRIES = 1  # Minimal retries - correct sample should work first time

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("warroom_ingestion.log", mode='a', encoding='utf-8')
    ]
)
logger = logging.getLogger("SmartClassifier")


# ============================================================
# PROMPTS
# ============================================================

QUICK_CLASSIFY_PROMPT = """Analizza questo documento finanziario.

{content_description}

Rispondi SOLO in JSON:
{{
  "document_type": "CRYPTO" | "STOCKS" | "COMMODITIES" | "CASH_MOVEMENTS" | "UNKNOWN",
  "content_summary": "breve descrizione del contenuto",
  "is_trading_data": true se contiene operazioni buy/sell di asset,
  "has_structured_tables": true (sempre per Excel/CSV),
  "estimated_rows": "few" | "tens" | "hundreds" | "many",
  "data_quality": "HIGH" | "MEDIUM" | "LOW",
  "should_process": true se utile per portafoglio,
  "skip_reason": "motivo se should_process=false, altrimenti null"
}}
"""

PHASE1_TEXT_PROMPT = """Analizza questo documento finanziario strutturato.

COLONNE: {columns}
PRIME {num_rows} RIGHE:
{sample_rows}

Rispondi in JSON:
{{
  "document_type": "CRYPTO" | "STOCKS" | "COMMODITIES" | "CASH_MOVEMENTS" | "UNKNOWN",
  "content_summary": "descrizione contenuto",
  "asset_type": "CRYPTO" | "STOCK" | "COMMODITY" | "CASH",
  "is_trading_data": true/false,
  "should_process": true se contiene dati utili per portafoglio (es. trade, dividendi), false se irrilevante (es. disclaimer, movimenti interni),
  "column_observations": {{
    "symbol_candidates": ["colonne con simboli"],
    "date_candidates": ["colonne con date"],
    "quantity_candidates": ["colonne con quantità"],
    "price_candidates": ["colonne con prezzi"],
    "amount_candidates": ["colonne con importi"]
  }}
}}
"""

PHASE2_TEXT_PROMPT = """Basandoti su questa analisi, mappa le colonne esatte.

ANALISI: {phase1}
COLONNE DISPONIBILI: {columns}

Rispondi SOLO in JSON:
{{
  "column_mapping": {{
    "symbol": "colonna esatta o null",
    "date": "colonna esatta o null",
    "operation_type": "colonna esatta o null",
    "quantity": "colonna esatta o null",
    "price": "colonna esatta o null",
    "total_amount": "colonna esatta o null",
    "fees": "colonna esatta o null",
    "currency": "colonna esatta o null",
    "balance": "colonna esatta o null"
  }},
  "date_format": "formato data",
  "number_format": "EU o US"
}}
"""

VISION_CLASSIFY_PROMPT = """Look at this financial document.

What type is it?
- CRYPTO: cryptocurrency (BTC, ETH, crypto symbols)
- STOCKS: stock trades (company tickers, shares)
- COMMODITIES: metals (XAU gold, XAG silver)
- CASH_MOVEMENTS: bank transfers, payments only

Reply ONLY with JSON:
{"document_type": "CRYPTO" or "STOCKS" or "COMMODITIES" or "CASH_MOVEMENTS", "content": "what the document shows", "rows": "few/tens/hundreds", "quality": "HIGH/MEDIUM/LOW", "process": true or false}
"""

VISION_MAPPING_PROMPT = """Look at this financial document image.

List the column headers you see. Answer ONLY with valid JSON:
{"columns": ["col1", "col2", ...], "symbol_column": "name or null", "date_column": "name or null", "amount_column": "name or null"}
"""


# ============================================================
# LOADERS
# ============================================================

def load_excel_data(filepath: Path, max_rows: int = QUICK_SAMPLE_ROWS) -> tuple:
    """Load Excel and return columns + sample."""
    import openpyxl
    import pandas as pd
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    if ws.max_column == 1:  # CSV-in-cell format
        csv_lines = [row[0] for row in ws.iter_rows(min_row=1, values_only=True) if row[0]]
        csv_text = '\n'.join(csv_lines)
        df = pd.read_csv(StringIO(csv_text))
    else:
        df = pd.read_excel(filepath)
    
    columns = list(df.columns)
    sample = df.head(max_rows).to_string(index=False)
    return columns, sample, len(df)


def load_csv_data(filepath: Path, max_rows: int = QUICK_SAMPLE_ROWS) -> tuple:
    """Load CSV and return columns + sample."""
    import pandas as pd
    
    for encoding in ['utf-8', 'latin-1', 'cp1252']:
        try:
            df = pd.read_csv(filepath, encoding=encoding, nrows=max_rows*2)
            break
        except:
            continue
    
    columns = list(df.columns)
    sample = df.head(max_rows).to_string(index=False)
    return columns, sample, len(df)


def pdf_to_images(filepath: Path, pages: list = None) -> list:
    """Convert PDF pages to base64 images. 
    STRATEGY: 2 Pages (First & Last).
    (6 images caused HTTP 400 Payload Too Large error on Native Ollama).
    """
    from pdf2image import convert_from_path, pdfinfo_from_path
    
    # Poppler path for Windows
    poppler_path = project_root / "tools" / "poppler-24.02.0" / "Library" / "bin"
    poppler_arg = str(poppler_path) if poppler_path.exists() else None
    
    try:
        # Get total page count first
        info = pdfinfo_from_path(str(filepath), poppler_path=poppler_arg)
        total_pages = info['Pages']
        
        # Determine pages to extract
        # STRATEGY: First 2 Pages only (Fast Classification)
        # Detailed extraction happens in Phase 3 (reading all pages)
        total_pages = info['Pages']
        pages_to_convert = [1]
        if total_pages >= 2:
            pages_to_convert.append(2)
            
        logger.info(f"   📄 Extracting {len(pages_to_convert)} pages for CLASSIFICATION: {pages_to_convert} (Total: {total_pages})")
        
        base64_images = []
        
        # Process page by page
        for page_num in pages_to_convert:
            images = convert_from_path(
                str(filepath),
                first_page=page_num,
                last_page=page_num,
                dpi=110, 
                poppler_path=poppler_arg
            )
            for img in images:
                img.thumbnail((1000, 1000)) 
                buffer = BytesIO()
                img.save(buffer, format='JPEG', quality=80) 
                b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
                base64_images.append(b64)
        
        return base64_images
    except Exception as e:
        logger.error(f"PDF to image conversion failed: {e}")
        return []


def call_vision_ollama(prompt: str, images: list, model: str = VISION_MODEL) -> str:
    """Call Ollama with vision model - optimized for stability."""
    import requests
    
    url = "http://localhost:11434/api/generate"
    
    payload = {
        "model": model,
        "prompt": prompt,
        "images": images,
        "stream": False,
        "keep_alive": "30m",  # Keep loaded longer to prevent SSD re-reads
        "options": {
            "temperature": 0.0,
            "num_predict": 500, # Shorter response is fine
            "num_ctx": 4096
        }
    }
    
    try:
        logger.info(f"[VISION] 📤 Sending {len(images)} image(s) to {model}...")
        # Increase timeout drastically for SSD loading (5 min)
        resp = requests.post(url, json=payload, timeout=300)
        
        if resp.status_code == 200:
            result = resp.json().get("response", "")
            snippet = result[:200].replace('\n', ' ')
            logger.info(f"[VISION] 📥 Received {len(result)} chars: {snippet}...")
            return result
        else:
            logger.error(f"[VISION] Error: {resp.status_code}")
            return ""
    except requests.exceptions.Timeout:
        logger.warning(f"[VISION] ⏳ Timeout (300s). Unloading model to recover resource...")
        # Don't unload immediately if we are in a batch, but here we failed.
        unload_vision_model()
        return ""
    except Exception as e:
        logger.error(f"[VISION] Exception: {e}")
        return ""


def unload_vision_model():
    """Unload vision model to free VRAM."""
    import requests
    try:
        requests.post("http://localhost:11434/api/generate", json={
            "model": VISION_MODEL,
            "keep_alive": 0
        }, timeout=10)
        logger.info("[VISION] 🧹 Model unloaded from VRAM")
    except:
        pass


def extract_json_from_text(text: str) -> str:
    """Robustly extract JSON from model output even with chatter."""
    try:
        start = text.find('{')
        end = text.rfind('}')
        if start != -1 and end != -1 and end > start:
            return text[start:end+1]
        return text
    except:
        return text


# ============================================================
# RETRY & VALIDATION
# ============================================================

def parse_with_retry(response: str, max_retries: int = MAX_RETRIES) -> dict:
    """Parse JSON with retry logic + Regex Fallback for Vision."""
    if not response:
        return None
    
    # 1. First try: standard JSON extraction
    clean_json = extract_json_from_text(response)
    try:
        result = json.loads(clean_json)
        return result
    except:
        pass
    
    # 2. Second try: garbage collector parser
    result = parse_json_garbage(response)
    if result:
        return result
        
    # 3. Third try: Regex fallback for Vision-style formatted text
    # Matches: **Key:** Value, Key: Value, etc.
    if "Type" in response or "TYPE" in response:
        logger.info("   ⚠️ JSON failed. Attempting Regex fallback...")
        fallback = {
            "document_type": "UNKNOWN",
            "is_trading_data": False,
            "should_process": False,
            "has_structured_tables": False,
            "data_quality": "LOW",
            "fallback_used": True
        }
        
        # Extract Content Summary and Date
        content_match = re.search(r"(?:\*\*|#)?\s*(?:Content|Summary)[:\s]+(?:\*\*)?\s*([^\n]+)", response, re.IGNORECASE)
        if content_match:
            fallback["content_summary"] = content_match.group(1).strip()
            
        dates = re.findall(r"20\d{2}", response)
        if dates:
            fallback["date_range_visible"] = f"{min(dates)}-{max(dates)}"
            
        return fallback

    if max_retries > 0:
        logger.warning(f"   ⚠️ Parse failed. Retrying...")
    
    return None


def validate_column_mapping(mapping: dict, actual_columns: list) -> dict:
    """Validate that mapped columns exist in the actual data."""
    if not mapping:
        return {}
    
    validated = {}
    for field, column in mapping.items():
        if column and column in actual_columns:
            validated[field] = column
        elif column:
            # Try case-insensitive match
            matches = [c for c in actual_columns if c.lower() == column.lower()]
            validated[field] = matches[0] if matches else None
        else:
            validated[field] = None
    
    return validated


# ============================================================
# CLASSIFICATION FUNCTIONS
# ============================================================

def quick_classify_text(filepath: Path) -> dict:
    """Quick classification for Excel/CSV files."""
    ext = filepath.suffix.lower()
    
    if ext in ['.xlsx', '.xls']:
        columns, sample, _ = load_excel_data(filepath, max_rows=5)
    else:
        columns, sample, _ = load_csv_data(filepath, max_rows=5)
    
    content = f"COLONNE: {', '.join(columns)}\nPRIME RIGHE:\n{sample}"
    prompt = QUICK_CLASSIFY_PROMPT.format(content_description=content)
    
    response = call_ollama(prompt, context="QUICK_CLASSIFY", model=QWEN_MODEL)
    return parse_with_retry(response) or {"error": "Quick classify failed"}


def quick_classify_vision(filepath: Path) -> dict:
    """Quick classification for PDF files using vision."""
    images = pdf_to_images(filepath, pages=[1])
    
    if not images:
        return {"error": "Could not convert PDF to image"}
    
    response = call_vision_ollama(VISION_CLASSIFY_PROMPT, images)
    return parse_with_retry(response) or {"error": "Vision classify failed"}


def full_classify_text(filepath: Path) -> dict:
    """Full two-phase classification for Excel/CSV."""
    ext = filepath.suffix.lower()
    
    if ext in ['.xlsx', '.xls']:
        columns, sample, row_count = load_excel_data(filepath, max_rows=FULL_SAMPLE_ROWS)
    else:
        columns, sample, row_count = load_csv_data(filepath, max_rows=FULL_SAMPLE_ROWS)
    
    # Phase 1: Understanding
    logger.info("   🧠 Phase 1: Understanding...")
    prompt1 = PHASE1_TEXT_PROMPT.format(
        columns=", ".join(columns),
        num_rows=min(row_count, FULL_SAMPLE_ROWS),
        sample_rows=sample
    )
    response1 = call_ollama(prompt1, context="PHASE1_TEXT", model=QWEN_MODEL)
    phase1 = parse_with_retry(response1)
    
    if not phase1:
        return {"error": "Phase 1 failed"}
    
    # Phase 2: Mapping
    logger.info("   🔧 Phase 2: Mapping...")
    prompt2 = PHASE2_TEXT_PROMPT.format(
        phase1=json.dumps(phase1, ensure_ascii=False),
        columns=", ".join(columns)
    )
    response2 = call_ollama(prompt2, context="PHASE2_TEXT", model=QWEN_MODEL)
    phase2 = parse_with_retry(response2)
    
    if not phase2:
        return {"error": "Phase 2 failed", "phase1": phase1}
    
    # Validate mapping
    if phase2.get("column_mapping"):
        phase2["column_mapping"] = validate_column_mapping(
            phase2["column_mapping"], columns
        )
    
    return {
        **phase1,
        **phase2,
        "_columns": columns,
        "_row_count": row_count
    }


def full_classify_pdf_text(filepath: Path, text_content: str) -> dict:
    """Classify PDF using extracted text (Fast Mode)."""
    # Truncate content for prompt
    preview = text_content[:2000]
    
    prompt = f"""Analyze this financial document text:
    
{preview}

Determine:
1. Document Type (STOCKS, CRYPTO, COMMODITIES, CASH_MOVEMENTS, ACCOUNT, UNKNOWN)
2. Asset Type (STOCK, ETF, CRYPTO, COMMODITIES, EUR, USD)
3. Should Process? (True if it contains individual trade executions or dividends. False if it's just a monthly balance summary or tax report without trade details.)

Reply ONLY JSON:
{{"document_type": "...", "asset_type": "...", "process": true/false, "reason": "..."}}
"""
    
    response = call_ollama(prompt, context=f"Classify PDF Text: {filepath.name}")
    data = parse_json_garbage(response) or {}
    
    should_process = data.get("process", False)
    
    # Filename Heuristics to override False negatives
    fname = filepath.name.lower()
    if "trading" in fname or "pnl" in fname:
        should_process = True
        
    return {
        "document_type": data.get("document_type", "UNKNOWN"),
        "asset_type": data.get("asset_type", "UNKNOWN"),
        "should_process": should_process,
        "confidence": "HIGH",
        "extraction_strategy": "TEXT_LLM",
        "strategy_reason": "Native PDF Text Layer",
        "content_summary": data.get("reason", "Text analysis"),
        "has_text_layer": True
    }


def full_classify_vision(filepath: Path) -> dict:
    """Full two-phase classification for PDF using sequential vision analysis."""
    # Get up to 6 images (Start/Mid/End)
    images = pdf_to_images(filepath)
    
    if not images:
        return {"error": "Could not convert PDF to images"}
    
    # Unload text model
    unload_text_model()
        
    logger.info(f"   👁️ Vision Classification (Sequential Analysis of {len(images)} pages)...")
    
    final_result = {
        "document_type": "UNKNOWN",
        "should_process": False,
        "content_summary": "",
        "confidence_score": 0.0
    }
    
    found_types = []
    summaries = []
    
    # SEQUENTIAL ANALYSIS LOOP
    for i, img in enumerate(images):
        logger.info(f"      📸 Analyzing Page {i+1}/{len(images)}...")
        # Send ONE image at a time to avoid payload limits
        response = call_vision_ollama(VISION_CLASSIFY_PROMPT, [img])
        partial_result = parse_with_retry(response)
        
        if partial_result:
            # Aggregate findings
            doc_type = partial_result.get("document_type", "UNKNOWN")
            should_proc = partial_result.get("should_process", False) # or key 'process' from vision prompt
            if "process" in partial_result: should_proc = partial_result["process"] # Handle alias
            
            if doc_type not in ["UNKNOWN", "CASH_MOVEMENTS"]:
                found_types.append(doc_type)
            
            if should_proc:
                final_result["should_process"] = True
                
            if "content" in partial_result:
                 summaries.append(f"P{i+1}: {partial_result['content']}")
            elif "content_summary" in partial_result:
                 summaries.append(f"P{i+1}: {partial_result['content_summary']}")

    # SYNTHESIS LOGIC
    if found_types:
        # Pick most frequent or first meaningful type
        from collections import Counter
        most_common = Counter(found_types).most_common(1)
        final_result["document_type"] = most_common[0][0]
    
    final_result["content_summary"] = " | ".join(summaries) if summaries else "No summary available"
    
    # --- FILENAME HEURISTIC POST-PROCESSING ---
    filename_lower = filepath.name.lower()
    if not final_result["should_process"]:
        if "crypto" in filename_lower:
            final_result["should_process"] = True
            final_result["document_type"] = "CRYPTO" 
            final_result["_heuristic_override"] = "Filename contains 'crypto'"
            logger.info(f"      📂 Filename heuristic forced CANDIDATE (crypto)")
        elif "trading" in filename_lower or "pnl" in filename_lower:
            final_result["should_process"] = True
            if final_result["document_type"] not in ["STOCKS", "CRYPTO"]:
                final_result["document_type"] = "STOCKS"
            final_result["_heuristic_override"] = "Filename contains 'trading/pnl'"
            logger.info(f"      📂 Filename heuristic forced CANDIDATE (trading)")
            
    # --- STRATEGY SELECTION ---
    # Check for text layer to recommend strategy
    has_text_layer = False
    text_len = 0
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(filepath))
        # Check first 2 pages
        for p in reader.pages[:2]:
            text_len += len(p.extract_text() or "")
        has_text_layer = text_len > 100
    except:
        pass
        
    final_result["has_text_layer"] = has_text_layer
    final_result["text_layer_len"] = text_len
    
    # DECISION MATRIX
    # If text layer is present -> Recommend TEXT_LLM (Fast, Safe)
    # If no text layer -> Recommend VISION_LLM (Slow, only way)
    if has_text_layer:
        final_result["extraction_strategy"] = "TEXT_LLM"
        final_result["strategy_reason"] = f"Native text layer detected ({text_len} chars)"
    else:
        final_result["extraction_strategy"] = "VISION_LLM"
        final_result["strategy_reason"] = "Scanned document or no selectable text"
    
    return {
        **final_result,
        "column_mapping": None, 
        "_pdf_note": "Sequential analysis complete"
    }


def unload_text_model():
    """Helper to unload text model."""
    import requests
    try:
        requests.post("http://localhost:11434/api/generate", json={"model": QWEN_MODEL, "keep_alive": 0})
        logger.info(f"   🧹 Unloaded Text Model ({QWEN_MODEL}) before Vision task")
    except:
        pass


# ============================================================
# MAIN ENTRY POINTS
# ============================================================

def quick_classify(filepath: Path) -> dict:
    """Quick classification based on file type."""
    filepath = Path(filepath)
    ext = filepath.suffix.lower()
    
    logger.info(f"⚡ Quick-classifying: {filepath.name}")
    
    if ext in ['.xlsx', '.xls', '.csv']:
        return quick_classify_text(filepath)
    elif ext == '.pdf':
        return quick_classify_vision(filepath)
    else:
        return {"error": f"Unsupported file type: {ext}"}


def classify_document(filepath: Path) -> dict:
    """Full classification based on file type."""
    filepath = Path(filepath)
    ext = filepath.suffix.lower()
    
    logger.info(f"📄 Full classification: {filepath.name}")
    
    if ext in ['.xlsx', '.xls', '.csv']:
        return full_classify_text(filepath)
    elif ext == '.pdf':
        # 1. OPTIMIZATION: Check for dense text layer first
        text_content = ""
        try:
            reader = PdfReader(str(filepath))
            # Sample first 2 pages
            for p in reader.pages[:2]:
                text_content += (p.extract_text() or "")
        except Exception as e:
            logger.warning(f"Failed to check text layer: {e}")
            
        if len(text_content) > 100:
            logger.info(f"🚀 Speed Optimization: Using Text Classify ({len(text_content)} chars)")
            return full_classify_pdf_text(filepath, text_content)
        else:
            return full_classify_vision(filepath)
    else:
        return {"error": f"Unsupported file type: {ext}"}


# CLI
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python smart_classifier.py <file> [--quick]")
        sys.exit(1)
    
    target = Path(sys.argv[1])
    quick_mode = "--quick" in sys.argv
    
    if target.is_file():
        if quick_mode:
            result = quick_classify(target)
        else:
            result = classify_document(target)
        
        print("\n" + "="*60)
        print("RESULT:")
        print("="*60)
        print(json.dumps(result, indent=2, ensure_ascii=False))
