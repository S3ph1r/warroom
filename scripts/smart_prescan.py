"""
Smart Prescan - Document Registry & Deduplication

Scans inbox folders, extracts broker from folder name, fingerprints content,
and manages the document registry to avoid duplicates and re-processing.
"""

import sys
import json
import hashlib
import logging
from pathlib import Path
from datetime import datetime
from io import StringIO

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

logging.basicConfig(level=logging.INFO, format='%(asctime)s | %(levelname)s | %(message)s')
logger = logging.getLogger("SmartPrescan")

# Registry file location
REGISTRY_PATH = project_root / "inbox_manifest.json"


def load_registry() -> dict:
    """Load existing registry or create new one."""
    if REGISTRY_PATH.exists():
        with open(REGISTRY_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {
        "last_scan": None,
        "brokers": {}
    }


def save_registry(registry: dict):
    """Save registry to disk."""
    registry["last_scan"] = datetime.now().isoformat()
    with open(REGISTRY_PATH, 'w', encoding='utf-8') as f:
        json.dump(registry, f, indent=2, ensure_ascii=False)
    logger.info(f"📝 Registry saved: {REGISTRY_PATH}")


def extract_broker(filepath: Path) -> str:
    """Extract broker name from folder path."""
    # Assuming structure: .../inbox/BROKER/file.xlsx
    parts = filepath.parts
    for i, part in enumerate(parts):
        if part.lower() == "inbox" and i + 1 < len(parts):
            return parts[i + 1].upper()
    # Fallback: use parent folder name
    return filepath.parent.name.upper()


def fingerprint_excel(filepath: Path, sample_rows: int = 20) -> str:
    """Generate content fingerprint for Excel file."""
    import openpyxl
    import pandas as pd
    
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Handle CSV-in-cell format
        if ws.max_column == 1:
            csv_lines = [row[0] for row in ws.iter_rows(min_row=1, values_only=True) if row[0]]
            csv_text = '\n'.join(csv_lines)
            df = pd.read_csv(StringIO(csv_text))
        else:
            df = pd.read_excel(filepath)
        
        # Take sample and normalize
        sample = df.head(sample_rows).to_string(index=False)
        normalized = ''.join(sample.lower().split())
        
        return hashlib.md5(normalized.encode()).hexdigest()[:12]
    except Exception as e:
        logger.warning(f"   ⚠️ Could not fingerprint {filepath.name}: {e}")
        return hashlib.md5(filepath.name.encode()).hexdigest()[:12]


def fingerprint_csv(filepath: Path, sample_rows: int = 20) -> str:
    """Generate content fingerprint for CSV file."""
    import pandas as pd
    
    try:
        df = pd.read_csv(filepath, nrows=sample_rows)
        sample = df.to_string(index=False)
        normalized = ''.join(sample.lower().split())
        return hashlib.md5(normalized.encode()).hexdigest()[:12]
    except:
        return hashlib.md5(filepath.name.encode()).hexdigest()[:12]


def fingerprint_pdf(filepath: Path, sample_pages: int = 2) -> str:
    """Generate content fingerprint for PDF file."""
    import pdfplumber
    
    try:
        text_parts = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages[:sample_pages]:
                text = page.extract_text(layout=False)
                if text:
                    text_parts.append(text)
        
        full_text = ''.join(text_parts)
        normalized = ''.join(full_text.lower().split())
        return hashlib.md5(normalized.encode()).hexdigest()[:12]
    except:
        return hashlib.md5(filepath.name.encode()).hexdigest()[:12]


def fingerprint_file(filepath: Path) -> str:
    """Generate content fingerprint based on file type."""
    ext = filepath.suffix.lower()
    
    if ext in ['.xlsx', '.xls']:
        return fingerprint_excel(filepath)
    elif ext == '.csv':
        return fingerprint_csv(filepath)
    elif ext == '.pdf':
        return fingerprint_pdf(filepath)
    else:
        return hashlib.md5(filepath.name.encode()).hexdigest()[:12]


def scan_broker_folder(folder: Path, registry: dict) -> list:
    """Scan a single broker's folder and return list of files to process."""
    broker = extract_broker(folder / "dummy.txt")  # Extract broker from path
    logger.info(f"\n📂 Scanning broker: {broker}")
    
    # Initialize broker in registry if needed
    if broker not in registry["brokers"]:
        registry["brokers"][broker] = {
            "source_folder": str(folder),
            "documents": {}
        }
    
    broker_docs = registry["brokers"][broker]["documents"]
    files_to_process = []
    
    # Scan all supported files
    all_files = list(folder.glob("*.xlsx")) + list(folder.glob("*.xls")) + \
                list(folder.glob("*.csv")) + list(folder.glob("*.pdf"))
    
    logger.info(f"   Found {len(all_files)} files")
    
    # Group by fingerprint to detect duplicates
    fingerprint_groups = {}
    
    for filepath in all_files:
        logger.info(f"   🔑 Fingerprinting: {filepath.name}")
        fp = fingerprint_file(filepath)
        
        if fp not in fingerprint_groups:
            fingerprint_groups[fp] = []
        fingerprint_groups[fp].append(filepath)
    
    # Process each fingerprint group
    for fp, files in fingerprint_groups.items():
        # Check if already processed
        if fp in broker_docs and broker_docs[fp].get("status") == "PROCESSED":
            logger.info(f"   ⏭️ Already processed: {files[0].name} (hash: {fp})")
            continue
        
        # If multiple files with same fingerprint, select preferred
        if len(files) > 1:
            logger.info(f"   🔄 Duplicate detected: {[f.name for f in files]}")
            preferred = select_preferred_format(files)
            logger.info(f"      → Preferred: {preferred.name}")
        else:
            preferred = files[0]
        
        files_to_process.append({
            "filepath": preferred,
            "broker": broker,
            "fingerprint": fp,
            "all_versions": [f.name for f in files]
        })
    
    return files_to_process


def select_preferred_format(files: list) -> Path:
    """Select preferred file format when duplicates exist."""
    # Priority: xlsx > xls > csv > pdf
    priority = {'.xlsx': 1, '.xls': 2, '.csv': 3, '.pdf': 4}
    
    sorted_files = sorted(files, key=lambda f: priority.get(f.suffix.lower(), 99))
    return sorted_files[0]


def scan_inbox(inbox_root: Path) -> list:
    """Scan entire inbox and return list of files to process."""
    registry = load_registry()
    all_files_to_process = []
    
    logger.info(f"🔍 Scanning inbox: {inbox_root}")
    
    # Check if inbox_root itself has files (Single Broker Mode)
    root_has_files = any(inbox_root.glob("*.xlsx")) or any(inbox_root.glob("*.csv")) or any(inbox_root.glob("*.pdf"))
    if root_has_files:
        logger.info(f"   📂 Detected Single Broker Mode (files in root)")
        files = scan_broker_folder(inbox_root, registry)
        all_files_to_process.extend(files)
    
    # Scan each broker subfolder
    for subfolder in inbox_root.iterdir():
        if subfolder.is_dir():
            # If we already scanned root, avoid scanning subfolders if they are just organizational (like PDF/)
            # But maybe we want recursive? 
            # Current logic: extract_broker uses folder name.
            # If root is REVOLUT, subfolder PDF would be broker PDF?
            # extract_broker uses "inbox/BROKER".
            
            # If we are in Single Broker Mode, subfolders might be parts of the same broker.
            # But scan_broker_folder re-extracts broker name.
            pass
            
            # Standard mode: scan subfolders as brokers
            if not root_has_files: 
                files = scan_broker_folder(subfolder, registry)
                all_files_to_process.extend(files)
            elif subfolder.name.upper() == "PDF": # Special handling for PDF subfolder in single mode
                # Treat as same broker?
                # scan_broker_folder will deduce broker from path.
                # if path is .../revolut/PDF, extract_broker might identify as PDF?
                # Let's trust scan_broker_folder logic but we need to ensure broker name matches.
                files = scan_broker_folder(subfolder, registry)
                all_files_to_process.extend(files)
    
    # Save updated registry
    save_registry(registry)
    
    logger.info(f"\n✅ Scan complete: {len(all_files_to_process)} files to process")
    
    return all_files_to_process


def mark_processed(filepath: Path, document_type: str, row_count: int = 0):
    """Mark a document as processed in the registry."""
    registry = load_registry()
    broker = extract_broker(filepath)
    fp = fingerprint_file(filepath)
    
    if broker not in registry["brokers"]:
        registry["brokers"][broker] = {"source_folder": str(filepath.parent), "documents": {}}
    
    registry["brokers"][broker]["documents"][fp] = {
        "files": [filepath.name],
        "document_type": document_type,
        "status": "PROCESSED",
        "processed_at": datetime.now().isoformat(),
        "row_count": row_count
    }
    
    save_registry(registry)


def mark_skipped(filepath: Path, reason: str):
    """Mark a document as skipped in the registry."""
    registry = load_registry()
    broker = extract_broker(filepath)
    fp = fingerprint_file(filepath)
    
    if broker not in registry["brokers"]:
        registry["brokers"][broker] = {"source_folder": str(filepath.parent), "documents": {}}
    
    registry["brokers"][broker]["documents"][fp] = {
        "files": [filepath.name],
        "status": "SKIPPED",
        "reason": reason,
        "skipped_at": datetime.now().isoformat()
    }
    
    save_registry(registry)


# CLI Interface
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python smart_prescan.py <inbox_folder>")
        sys.exit(1)
    
    inbox = Path(sys.argv[1])
    
    if inbox.is_dir():
        files = scan_inbox(inbox)
        print("\n" + "="*60)
        print("FILES TO PROCESS:")
        print("="*60)
        for f in files:
            print(f"  [{f['broker']}] {f['filepath'].name} (fp: {f['fingerprint']})")
    else:
        print(f"Error: {inbox} is not a directory")
        sys.exit(1)
